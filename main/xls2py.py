# encoding=utf-8

import os
import logging, pprint

import openpyxl as pyxl
from grammer_checkers.type_check import type_checker
from grammer_checkers.value_check import val_checker
import config, inner_config
import util
from .dump_xlspy import dump_dict_to_file

class XlsConverter(object):

    def __init__(self, file_name, output, style):
        self.file_name = file_name

        self.output = str(output)
        self.style = style
        self.proper_infos = {}
        self.logger = None
        self.cur_sheet_name = ""

        self.workbook = None
        self.errors = []
        self.output_dict = {}

    def convert(self):

        self.xls2py()

    def xls2py(self):

        self.workbook = pyxl.load_workbook(filename=self.file_name, read_only=True)
        sheet_names = self.workbook.sheetnames
        if config.EXPORT_LIST_SHEET not in sheet_names:
            self.add_error_log("%s not in sheet_names" % config.EXPORT_LIST_SHEET)
            return
        export_ws = self.workbook.get_sheet_by_name(config.EXPORT_LIST_SHEET)
        export_sheet_names = self.get_export_sheet_names(export_ws)
        self.get_logger().info("sheet_names: %s", export_sheet_names)
        for sheet_name in export_sheet_names:
            sheet_dict = self.xls2dict(sheet_name)
            self.output_dict[sheet_name] = sheet_dict
        self.gen_output()

    def get_export_sheet_names(self, export_ws):

        max_row_num = export_ws.max_row
        sheet_names = []
        for idx in range(config.PRO_MAX_ROW+1, max_row_num+1):
            name = export_ws[idx][0]
            if not name.value:
                error_info = "No Sheet Name in row={row}, sheet_names should be together in sheet={sheet}".format(\
                        row=idx+1, sheet=config.EXPORT_LIST_SHEET)
                self.add_error_log(error_info)
                continue
            sheet_names.append(name.value)
        return sheet_names

    def xls2dict(self, sheet_name):
        self.cur_sheet_name = sheet_name
        errors = self.init_proper_info(sheet_name)
        if errors:
            print_error_contents(errors)
            return {}

        sheet = self.workbook[sheet_name]
        max_row_num = sheet.max_row
        sheet_proper_info = self.proper_infos[sheet_name]
        sheet_cfg = sheet_proper_info["config"]
        max_col_num = sheet_cfg["max_col_num"]
        pro_list = sheet_proper_info["pro_list"]
        output_dict = {}
        for rdx, cur_row in enumerate(list(sheet.rows)):
            if rdx < config.PRO_MAX_ROW:
                continue
            cur_dict = {}
            if not cur_row[0].value:
                continue
            for cdx in range(0, max_col_num):
                pro_info = pro_list[cdx]
                type_info = pro_info["type"]
                limit_cfg = pro_info["limit"]

                cell = cur_row[cdx]
                tbl_value = None
                type_name = limit_cfg["info"][0]
                if not cell.value:
                    if type_name in [type_checker.EXISTS, type_checker.UNIQUE]:
                        self.add_error_log("No value for pro={pro} in row={row}".format( \
                            pro=pro_info["name"], row=rdx+1))
                        continue
                    elif limit_cfg["info"][0] == type_checker.DEFAULT:
                        tbl_value = limit_cfg["def_val"]
                    else:
                        continue

                else:
                    tbl_value = val_checker.get_real_val(cell.value, type_info, self.get_logger())
                    if tbl_value is None:
                        self.add_error_log("Value intepret error: pro={pro}, rdx={row}".format(\
                            pro=pro_info["name"], row=rdx+1))

                    if type_name in [type_checker.UNIQUE]:
                        if tbl_value in limit_cfg["uniq"]:
                            self.add_error_log("repeated value of value={val}, origin in " \
                                "row={row}".format(val=tbl_value, row=limit_cfg["uniq"]))
                            continue
                        limit_cfg["uniq"][tbl_value] = rdx
                    elif type_name in [type_checker.STAR]:
                        pass
                    elif type_name not in [type_checker.DEFAULT, type_checker.EXISTS]:
                        self.add_error_log("Invalid limit symbol({symbol}) for pro={pro}".format(
                            pro=pro_info["name"], symbol=limit_cfg["info"][0]))

                if cdx == 0:
                    output_dict[tbl_value] = {}
                    cur_dict = output_dict[tbl_value]
                elif tbl_value is not None:
                    cur_dict[pro_info["name"]] = tbl_value

        return output_dict

    def init_proper_info(self, sheet_name):
        sheet = self.workbook[sheet_name]
        proper_infos = {
            "pro_list": [],
            "config": {},
        }
        self.proper_infos[sheet_name] = proper_infos
        name_row = sheet[config.PRO_NAME_ROW]
        type_row = sheet[config.TYPE_ROW]
        limit_row = sheet[config.LIMIT_ROW]

        tmp_pro_names = {}

        max_col_num = query_max_col(name_row)
        proper_infos["config"]["max_col_num"] = max_col_num
        for idx in range(0, max_col_num):
            pro_name = name_row[idx].value
            if pro_name.startswith(config.COMMENT_PRO_PREFIX):
                # 被注释掉的列，不参与导表
                self.get_logger().info("pro_name={pro} startswith {prefix}, ignored".format(\
                    pro=pro_name, prefix=config.COMMENT_PRO_PREFIX))
                continue
            if not pro_name:
                error_info = "No property name in col={col}".format(col=idx+1)
                self.add_error_log(error_info)
                continue
            elif pro_name in tmp_pro_names:
                ori_idx = tmp_pro_names[pro_name]
                error_info = "Property={pro} repeated," \
                    " origin in col={col}".format(pro=pro_name, col=ori_idx)
                self.add_error_log(error_info)
                continue
            tmp_pro_names[pro_name] = idx+1

            proper_info = {}
            proper_info["name"] = pro_name

            type_str = type_row[idx].value
            if not type_str:
                error_info = "No property type for" \
                    "{pro} in col={col}".format(pro=pro_name, col=idx+1)
                self.add_error_log(error_info)
            type_info = type_checker.get_type_info(type_str, self.get_logger())
            proper_info["type"] = type_info

            limit_str = limit_row[idx].value
            if not limit_str:
                error_info = "No limit for {pro} in col={col}".format(pro=pro_name, col=idx+1)
                self.add_error_log(error_info)
            limit_cfg = {
                "info": "",
                "def_val": None,
                "uniq": {},
            }
            limit_cfg["info"] = type_checker.get_limit_info(limit_str)
            limit_type = limit_cfg["info"][0]
            if limit_type == type_checker.DEFAULT:
                def_val = val_checker.get_real_val(limit_cfg["info"][1], type_info, self.get_logger())
                if def_val is None:
                    self.add_error_log("Value Error: pro_name={pro}, default_value={def_val} " \
                        "not match type={typ}".format(\
                        pro=pro_name, def_val=def_val, typ=type_info[0]))
                else:
                    limit_cfg["def_val"] = def_val

            proper_info["limit"] = limit_cfg
            proper_infos["pro_list"].append(proper_info)

        # print(sheet_name, '\n'.join(str(x) for x in self.proper_infos[sheet_name]["pro_list"]))

    def add_error_log(self, error_info):
        # self.errors.append(error_info)
        prefix = "file: {file}, sheet: {sheet}".format( \
            file=self.file_name, sheet=self.cur_sheet_name)
        self.get_logger().error(prefix + " " + error_info)

    def gen_output(self):
        # output = self.output
        dump_dict_to_file(self.output_dict, self.output, \
            self.style, self.get_logger(), True)
        self.check_dump_res()

    def check_dump_res(self):

        py_name = self.output.split(os.path.sep)[-1]
        for k in self.output_dict:
            # exec("import {src_dir}.{name}.tbl_{tbl}".\
            #     format(src_dir=config.SRC_DIR, name=py_name, tbl=k))
            tbl_name = util.get_table_name(py_name, k)
            exec("from {src_dir} import {name};{name}.{tbl}".\
                format(src_dir=config.SRC_DIR, name=py_name, tbl=tbl_name))
            exec("from {src_dir}.{name} import {tbl}".\
                format(src_dir=config.SRC_DIR, name=py_name, tbl=tbl_name))

    def get_logger(self):
        if not self.logger:
            self.logger = util.create_logger(self.file_name)
        return self.logger

def query_max_col(row):
    max_col_num = 1
    for idx, cell in enumerate(row):
        if not isinstance(cell.value, type(None)):
            if idx > max_col_num:
                max_col_num = idx
    return max_col_num

def print_error_contents(errors):
    print('\n'.join(errors))
